# main_pipeline.py
# Extracts 10 most relevant comments — 3-4 per component
# Focused on quality issues only, slow Gemini calls to avoid rate limit

import os, io, re, json, time, random, hashlib, warnings
import numpy as np
import pandas as pd
warnings.filterwarnings("ignore")

from googleapiclient.discovery import build
import google.generativeai as genai
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from gdrive_sync import download_from_onedrive, upload_to_onedrive

# ── KEYS ─────────────────────────────────────────────────────────
YOUTUBE_API_KEY = os.getenv("YOUTUBE_API_KEY", "")
GEMINI_API_KEY  = os.getenv("GEMINI_API_KEY",  "")
LOCAL_FILE      = "FINAL_STRUCTURED_COMPLAINTS.xlsx"
LOG_FILE        = "yt_last_run.json"

# ── CAP — 10 total, balanced across 3 components ─────────────────
MAX_TOTAL          = 10   # total comments per run
MAX_PER_GROUP      = 4    # max per component (EV/Hybrid/PBD)
MAX_PER_BUCKET     = 2    # max per problem type — ensures diversity
MIN_WORDS          = 8
MAX_WORDS          = 150

# ── APIS ─────────────────────────────────────────────────────────
youtube       = build("youtube", "v3", developerKey=YOUTUBE_API_KEY)
genai.configure(api_key=GEMINI_API_KEY)
analyzer      = SentimentIntensityAnalyzer()

analyzer.lexicon.update({
    "breakdown":-3.0,"broke":-2.5,"failed":-2.5,"failure":-2.5,
    "defect":-2.5,"fault":-2.0,"malfunction":-2.5,"problem":-2.0,
    "issue":-1.8,"drain":-2.0,"stuck":-2.5,"rattle":-2.0,
    "recall":-3.0,"dangerous":-3.0,"frustrated":-2.5,
    "terrible":-3.0,"worst":-3.0,"avoid":-2.0,"poor quality":-2.5,
    "smooth":2.0,"excellent":2.5,"amazing":2.5,"reliable":2.0,
    "recommend":2.0,"efficient":1.8,"satisfied":2.0,"best":2.5,
})

# ════════════════════════════════════════════════════════════════
# KEYWORDS
# ════════════════════════════════════════════════════════════════
KEYWORD_GROUPS = {
    "EV System": [
        "EV system","EV battery","battery drain","battery degradation",
        "battery dead","battery not charging","charge not working",
        "EV not charging","charging stopped","charging failed",
        "charging port","fast charging","DC charging","AC charging",
        "regenerative braking","electric motor failure","electric range",
        "range anxiety","range dropped","EV failure","state of charge",
        "electric vehicle","BMS","OBC","SOC","V2L",
        "Nexon EV","Tata EV","Punch EV","Tigor EV","MG ZS EV",
        "Windsor EV","Comet EV","Creta EV","Ioniq 5","Ioniq 6",
        "Mahindra EV","XUV400 EV","BYD Atto","BYD Seal","XEV 9e",
        "km on full charge","range on full charge","ev breakdown",
        "charging station problem","charge nahi ho raha",
        "battery khatam","range kam ho gayi",
    ],
    "Series Hybrid EV": [
        "series hybrid","hybrid electric vehicle","mild hybrid",
        "strong hybrid","full hybrid","self charging hybrid",
        "plug-in hybrid","hybrid mode","hybrid system",
        "hybrid system fault","hybrid battery issue",
        "hybrid not working","hybrid mileage","e-CVT","PHEV","MHEV",
        "Grand Vitara hybrid","Hyryder hybrid","Innova Hycross",
        "City hybrid","Invicto hybrid","Toyota hybrid","Maruti hybrid",
        "hybrid lag","e-CVT problem","hybrid jerky",
        "hybrid kitna deta hai","hybrid mode nahi aata",
    ],
    "Power Back Door": [
        "power back door","power tailgate","electric tailgate",
        "auto tailgate","hands free tailgate","power liftgate",
        "electric boot","automatic boot","smart tailgate",
        "kick sensor tailgate","foot sensor boot",
        "tailgate not opening","tailgate not closing","tailgate stuck",
        "tailgate sensor","tailgate malfunction","tailgate rattle",
        "tailgate noise","boot not closing","rear door motor",
        "liftgate motor","hands free boot","motorized tailgate",
        "electric dicky","power dicky","boot sensor",
        "foot gesture boot","dicky problem","dicky not opening",
        "boot khul nahi raha","dicky band nahi hoti",
    ]
}

OWNERS = {
    "EV System":        ("Nandini Sharma", "Nandinisharma8862@gmail.com"),
    "Series Hybrid EV": ("Nandini Singh",  "065038@fsm.ac.in"),
    "Power Back Door":  ("Tejas",          "Personaltej2909@gmail.com"),
    "Other":            ("General",        "general@maruti.com"),
}

PROBLEM_BUCKETS = {
    "battery_drain":    ["battery drain","battery dead","battery khatam",
                         "overnight","parasitic"],
    "charging_failure": ["charge not working","not charging","charging stopped",
                         "charging failed","charger not working","dc charging"],
    "range_issue":      ["range anxiety","range dropped","range problem",
                         "km on full charge","range kam","full charge"],
    "hybrid_mileage":   ["mileage","fuel efficiency","mileage kam",
                         "real world mileage","highway mileage"],
    "hybrid_system":    ["hybrid system","hybrid not working","hybrid lag",
                         "hybrid jerky","e-cvt","ecvt","hybrid fault"],
    "tailgate_sensor":  ["tailgate sensor","sensor fail","kick sensor",
                         "foot sensor","sensor not working","anti pinch"],
    "tailgate_stuck":   ["not opening","not closing","boot stuck",
                         "tailgate stuck","dicky nahi","band nahi"],
    "tailgate_noise":   ["tailgate rattle","tailgate noise","rattle",
                         "vibration","dicky noise"],
    "ev_breakdown":     ["breakdown","stalled","highway","towed",
                         "band ho gaya","stranded","broke down"],
    "ev_software":      ["software","ota","firmware","update","bug"],
}

# India-specific searches — tightly targeted
SEARCH_QUERIES = [
    # EV System — quality focused
    "Nexon EV battery problem India owner honest review",
    "Tata EV battery degradation India real owner experience",
    "EV charging failure India owner complaint 2024",
    "Nexon EV breakdown highway India owner review",
    "Tata Punch EV real range problem India honest",
    # Series Hybrid — quality focused
    "Grand Vitara hybrid problem India owner honest review",
    "Innova Hycross hybrid issue fault India owner",
    "Toyota hybrid system problem India real review",
    # Power Back Door — quality focused
    "power tailgate problem India SUV owner complaint",
    "electric tailgate sensor fail India car owner review",
    "electric dicky malfunction India SUV honest owner",
]

# COMPLAINT WORDS — expanded to catch hybrid/tailgate language
COMPLAINT_WORDS = [
    # English complaints
    "problem","issue","fault","defect","not working","failed","failure",
    "error","broken","complaint","bad","poor","worst","terrible","horrible",
    "stopped","not opening","not closing","drain","noise","rattle",
    "stuck","malfunction","breakdown","recall","repair","replace",
    "warning","disappointed","frustrated","pathetic","useless","avoid",
    "not good","disappointing","worst decision","regret","overheating",
    "overpriced","quality issue","quality problem","poor quality",
    "cheap quality","bad quality","build quality",
    # Hybrid specific
    "jerky","lag","vibration","shudder","judder","hesitation",
    "mileage dropped","fuel economy","less mileage","bad mileage",
    # Tailgate specific
    "not responding","sensor issue","not detecting","false trigger",
    "randomly opening","randomly closing","slow","too fast",
    # Hinglish
    "nahi chal raha","band ho gaya","kharab","dikkat","nahi ho raha",
    "nahi deta","nahi khul raha","problem aa raha","galat","bekar",
    "waste","paisa barbaad","thik nahi","kaam nahi karta",
]

INDIA_CONTEXT = [
    "india","indian","delhi","mumbai","bangalore","bengaluru","chennai",
    "hyderabad","pune","kolkata","maruti","tata","mahindra","nexon",
    "hyryder","vitara","hycross","innova","city hybrid","creta ev",
    "mg zs","windsor ev","xuv400","rupee","lakh","kmpl",
    "service centre","service center","dealer","showroom","emi",
    "punch ev","tigor","brezza","grand vitara","toyota","honda city",
    "invicto","suzuki","maruti suzuki",
]

COMPONENT_MUST_HAVE = {
    "EV System": [
        "battery","charging","charge","range","ev","electric",
        "bms","soc","motor","kwh","plug","charger","km",
    ],
    "Series Hybrid EV": [
        "hybrid","mileage","fuel","petrol","generator","e-cvt","ecvt",
        "self charging","strong hybrid","electric mode","kmpl","km",
    ],
    "Power Back Door": [
        "tailgate","boot","dicky","door","liftgate","trunk",
        "sensor","kick","hands free","motorized","electric boot",
        "power boot","auto close",
    ]
}

SPAM_PATTERNS = [
    r"^(nice|great|good|wow|amazing|superb|👍|❤️|🔥|lol)\s*[!.]*$",
    r"first\s*(comment|view)",
    r"(subscribe|like).{0,20}(channel|karo|please)",
    r"check.{0,10}(my|out).{0,20}channel",
    r"^\d+\s*(likes?|views?)",
    r"^[\s\W\d]{0,5}$",
]

REQUIRES_BOUNDARY = {"EV","BMS","OBC","SOC","V2L","PHEV","MHEV"}

# ════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════

def clean_text(text):
    t = str(text).lower()
    t = re.sub(r'[^a-z0-9 ]', '', t)
    return " ".join(t.split())


def comment_matches_keyword(text):
    if not text or len(text) < 20:
        return False, None
    tl = text.lower()
    for group, keywords in KEYWORD_GROUPS.items():
        for kw in keywords:
            if kw.upper() in REQUIRES_BOUNDARY:
                if re.search(r'\b' + re.escape(kw) + r'\b',
                             text, re.IGNORECASE):
                    return True, group
            else:
                if kw.lower() in tl:
                    return True, group
    return False, None


def get_problem_bucket(text):
    tl = text.lower()
    for bucket, signals in PROBLEM_BUCKETS.items():
        if any(s in tl for s in signals):
            return bucket
    return "general_complaint"


def score_comment(c):
    """
    Scores 0-100. Higher = more useful quality complaint.
    Prioritises: specific model + technical detail + likes + complaint specificity
    """
    text  = c.get("text", "")
    likes = int(c.get("likes", 0))
    words = len(text.split())
    tl    = text.lower()
    score = 0

    # Length — 15-80 words is ideal
    if 15 <= words <= 80:
        score += 25
    elif 8 <= words <= 120:
        score += 15
    else:
        score += 5

    # Likes from real owners
    if likes >= 20:   score += 25
    elif likes >= 10: score += 20
    elif likes >= 5:  score += 15
    elif likes >= 2:  score += 10
    elif likes >= 1:  score += 5

    # Specific Indian model mentioned
    indian_models = [
        "nexon ev","punch ev","tigor ev","windsor ev","comet ev",
        "creta ev","xuv400","xev 9e","grand vitara","hyryder",
        "innova hycross","city hybrid","invicto","mahindra ev",
    ]
    if any(m in tl for m in indian_models):
        score += 20

    # Technical quality terms
    quality_terms = [
        "bms","soc","kwh","kmpl","service centre","warranty","dealer",
        "ota","firmware","sensor","motor","e-cvt","range km","build",
        "quality","component","part","replaced","repair cost",
    ]
    score += min(15, sum(3 for t in quality_terms if t in tl))

    # Specific quality failure words
    quality_failures = [
        "not working","failed","breakdown","stuck","malfunction",
        "fault","defect","error","recall","replaced","repair",
        "quality issue","poor quality","bad quality","cheap",
    ]
    if any(qf in tl for qf in quality_failures):
        score += 10

    # Penalty — vague positive comments
    vague_positive = ["love it","great car","best car","awesome","superb"]
    if any(v in tl for v in vague_positive):
        score -= 15

    # Penalty — no India context
    if not any(ic in tl for ic in INDIA_CONTEXT):
        score -= 20

    return max(0, score)


def passes_filter(c, seen_clean, old_texts):
    text  = c.get("text", "").strip()
    group = c.get("keyword_group", "")
    words = text.split()

    if len(words) < MIN_WORDS:
        return False, "too_short"
    if len(words) > MAX_WORDS:
        return False, "too_long"

    for p in SPAM_PATTERNS:
        if re.search(p, text, re.IGNORECASE):
            return False, "spam"

    clean = clean_text(text)
    if clean in seen_clean or clean in old_texts:
        return False, "duplicate"

    tl = text.lower()

    # Must have at least 1 complaint word
    if not any(cw in tl for cw in COMPLAINT_WORDS):
        return False, "no_complaint"

    # Must have India context
    if not any(ic in tl for ic in INDIA_CONTEXT):
        return False, "no_india_context"

    # Must have component word
    must = COMPONENT_MUST_HAVE.get(group, [])
    if must and not any(kw in tl for kw in must):
        return False, "no_component_word"

    return True, "passed"


def select_top_10(filtered):
    """
    Selects top 10 from filtered comments.
    Balanced: max 4 per group, max 2 per problem bucket.
    Sorted by relevance score — highest first.
    """
    # Score all
    for c in filtered:
        c["_score"]  = score_comment(c)
        c["_bucket"] = get_problem_bucket(c.get("text", ""))

    # Sort by score
    sorted_c      = sorted(filtered, key=lambda x: x["_score"], reverse=True)
    selected      = []
    group_counts  = {}
    bucket_counts = {}

    for c in sorted_c:
        if len(selected) >= MAX_TOTAL:
            break
        group  = c.get("keyword_group", "Other")
        bucket = c.get("_bucket", "general_complaint")

        if group_counts.get(group, 0)  >= MAX_PER_GROUP:  continue
        if bucket_counts.get(bucket,0) >= MAX_PER_BUCKET: continue

        selected.append(c)
        group_counts[group]    = group_counts.get(group, 0) + 1
        bucket_counts[bucket]  = bucket_counts.get(bucket, 0) + 1

    print(f"\nSelected {len(selected)} from {len(filtered)} filtered:")
    print("  By component:")
    for g, n in sorted(group_counts.items()):
        print(f"    {g:<25}: {n}")
    print("  By problem type:")
    for b, n in sorted(bucket_counts.items(), key=lambda x: -x[1]):
        print(f"    {b:<25}: {n}")
    if selected:
        print(f"  Score range: "
              f"{selected[-1]['_score']} to {selected[0]['_score']}")
    return selected


def get_sentiment(text):
    if not text or len(str(text).strip()) < 5:
        return "Neutral", 0.0
    sentences = [s.strip() for s in
                 str(text).replace('\n', '. ').split('.')
                 if len(s.strip()) > 10]
    if len(sentences) > 1:
        scores   = [analyzer.polarity_scores(s) for s in sentences[:10]]
        compound = float(np.mean([s['compound'] for s in scores]))
    else:
        compound = analyzer.polarity_scores(text)['compound']
    label = ("Positive" if compound >=  0.15 else
             "Negative" if compound <= -0.10 else "Neutral")
    return label, round(compound, 4)


def get_next_sno(df):
    if df.empty or "S.No" not in df.columns:
        return 1
    nums = pd.to_numeric(df["S.No"], errors="coerce").dropna()
    return int(nums.max()) + 1 if not nums.empty else 1


def save_last_run():
    from datetime import datetime
    with open(LOG_FILE, "w") as f:
        json.dump({"last_run": datetime.now().isoformat()}, f)


# ════════════════════════════════════════════════════════════════
# YOUTUBE
# ════════════════════════════════════════════════════════════════

def search_videos(query, max_results=5):
    try:
        res = youtube.search().list(
            part="snippet", q=query, type="video",
            maxResults=max_results, regionCode="IN",
            relevanceLanguage="en", order="relevance",
            publishedAfter="2020-01-01T00:00:00Z"
        ).execute()
        return [
            {
                "video_id": item["id"]["videoId"],
                "title":    item["snippet"]["title"],
                "url":      f"https://youtube.com/watch?v={item['id']['videoId']}"
            }
            for item in res.get("items", [])
        ]
    except Exception as e:
        print(f"  Search error: {str(e)[:60]}")
        return []


def get_comments(video):
    comments = []
    try:
        req   = youtube.commentThreads().list(
            part="snippet", videoId=video["video_id"],
            maxResults=100, textFormat="plainText",
            order="relevance"
        )
        pages = 0
        while req and pages < 2:
            res   = req.execute()
            pages += 1
            for item in res.get("items", []):
                c    = item["snippet"]["topLevelComment"]["snippet"]
                text = c["textDisplay"].strip()
                matched, group = comment_matches_keyword(text)
                if matched:
                    comments.append({
                        "video_id":      video["video_id"],
                        "text":          text,
                        "date":          c["publishedAt"][:10],
                        "likes":         int(c.get("likeCount", 0)),
                        "video_title":   video["title"],
                        "video_url":     video["url"],
                        "keyword_group": group,
                    })
            req = youtube.commentThreads().list_next(req, res)
    except Exception as e:
        err = str(e)
        if "disabled" not in err.lower() and "403" not in err:
            print(f"  Comment error: {err[:60]}")
    return comments


# ════════════════════════════════════════════════════════════════
# GEMINI — with rate limit handling
# ════════════════════════════════════════════════════════════════

GEMINI_PROMPT = '''
You are an automotive quality analyst at Maruti Suzuki India.
Analyse this YouTube comment. Extract quality defect information only.
Focus on: what went wrong, why it went wrong, what the impact was.
Write "Not specified" for unknown fields.
Reply ONLY with valid JSON — absolutely no markdown or backticks.

Comment: "{comment}"
Video: "{title}"
Feature: "{group}"

{{
  "is_useful": true if this is a genuine quality complaint from a real owner. false if opinion/general query/spam.,
  "system_technology": "Specific e.g. EV System (Battery Drain) or EV System (Charging Failure) or Series Hybrid EV (Mileage Drop) or Series Hybrid EV (System Fault) or Power Back Door (Sensor Failure) or Power Back Door (Motor Failure)",
  "fn_type": "FN1 if component is broken or malfunctioning. FN2 if owner feedback or feature opinion.",
  "month": "Month if mentioned e.g. March. Not specified if not mentioned.",
  "year": "Year if mentioned e.g. 2024. Not specified if not mentioned.",
  "model": "Exact car model e.g. Tata Nexon EV 2023 or Maruti Grand Vitara Strong Hybrid. Not specified if not mentioned.",
  "defect_summary": "Professional one sentence: what quality issue did the owner face?",
  "cause": "Root cause if mentioned: BMS failure, sensor malfunction, software bug, design flaw, manufacturing defect. Not specified if unknown.",
  "action": "Resolution if mentioned: service centre visit, software update, part replacement, no fix found. Not specified if not mentioned.",
  "sentiment": "Negative, Positive, or Neutral"
}}
'''


def call_gemini_safe(c):
    """
    Calls Gemini with proper rate limit handling.
    Free tier = 15 requests/minute = 4 second gap minimum.
    We use 5 second gap to be safe.
    """
    prompt = GEMINI_PROMPT.format(
        comment=c.get("text", "")[:500],
        title=c.get("video_title", "")[:80],
        group=c.get("keyword_group", "")
    )

   for attempt in range(3):
    try:
        # Wait before each call to respect rate limit
        time.sleep(5)

        # Create Gemini model
        model = genai.GenerativeModel("gemini-2.0-flash")

        # Generate response
        resp = model.generate_content(
            prompt,
            request_options={"timeout": 120}
        )

        # Clean response
        raw = resp.text.strip()
        raw = raw.replace("```json", "").replace("```", "").strip()

        # Extract JSON safely
        s = raw.find("{")
        e = raw.rfind("}") + 1

        if s != -1 and e > s:
            raw = raw[s:e]

        # Return parsed JSON
        return json.loads(raw)

    except json.JSONDecodeError as je:
        print(f"JSON parse error: {str(je)[:40]}")
        time.sleep(3)

    except Exception as ex:
        err = str(ex)

        if "429" in err or "quota" in err.lower() or "rate" in err.lower():
            print(f"Rate limit → waiting 60s (attempt {attempt + 1})")
            time.sleep(60)

        elif "500" in err or "503" in err:
            print("Server error → waiting 15s")
            time.sleep(15)

        else:
            print(f"Gemini error: {err[:60]}")
            time.sleep(5)


    return None  # all 3 attempts failed


def fallback(c):
    return {
        "is_useful":        True,
        "system_technology": c.get("keyword_group", "Not specified"),
        "fn_type":          "Not specified",
        "month":            "Not specified",
        "year":             "Not specified",
        "model":            "Not specified",
        "defect_summary":   c.get("text", "")[:200],
        "cause":            "Not specified",
        "action":           "Not specified",
        "sentiment":        "Neutral",
    }


# ════════════════════════════════════════════════════════════════
# EXCEL
# ════════════════════════════════════════════════════════════════

EXCEL_COLUMNS = [
    "S.No",
    "System / Technology",
    "FN1 / FN2",
    "Month",
    "Year",
    "Model",
    "Defect / Feedback Summary",
    "Cause",
    "Action",
    "Sentiment",
    "VADER Score",
    "Owner Name",
    "Owner Email",
    "Status",
    "Source",
    "Video URL",
    "Date",
    "Original Comment",
]

COL_WIDTHS = [6, 28, 10, 12, 8, 28, 55, 40, 35, 12, 12, 18, 30, 10, 10, 45, 12, 60]


def save_excel(final_df, new_start):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import openpyxl.utils

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Complaints"

    hfill = PatternFill("solid", fgColor="1F3864")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    for ci, col in enumerate(EXCEL_COLUMNS, 1):
        cell           = ws.cell(row=1, column=ci, value=col)
        cell.fill      = hfill
        cell.font      = hfont
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(ci)
        ].width = w

    new_fill = PatternFill("solid", fgColor="D6EAF8")
    neg_fill = PatternFill("solid", fgColor="FFCCCC")
    pos_fill = PatternFill("solid", fgColor="CCFFCC")
    neu_fill = PatternFill("solid", fgColor="FFFACC")
    thin     = Side(style="thin", color="DDDDDD")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ri, row in final_df.iterrows():
        rn        = ri + 2
        is_new    = ri >= new_start
        sentiment = str(row.get("Sentiment", "")).lower()
        fill      = (new_fill if is_new else
                     neg_fill if sentiment == "negative" else
                     pos_fill if sentiment == "positive" else
                     neu_fill)

        values = [
            row.get("S.No", ""),
            row.get("System / Technology", ""),
            row.get("FN1 / FN2", ""),
            row.get("Month", ""),
            row.get("Year", ""),
            row.get("Model", ""),
            row.get("Defect / Feedback Summary", ""),
            row.get("Cause", ""),
            row.get("Action", ""),
            row.get("Sentiment", ""),
            row.get("VADER Score", ""),
            row.get("Owner Name", ""),
            row.get("Owner Email", ""),
            row.get("Status", "Open"),
            row.get("Source", "YouTube"),
            row.get("Video URL", ""),
            row.get("Date", ""),
            row.get("Original Comment", ""),
        ]

        for ci, val in enumerate(values, 1):
            cell           = ws.cell(row=rn, column=ci, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(LOCAL_FILE)


# ════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("PIPELINE STARTED")
    print(f"Target: {MAX_TOTAL} best quality complaints "
          f"({MAX_PER_GROUP} max per component)")
    print("=" * 60)

    # 1 — Download
    print("\n[1/7] Downloading from Google Drive...")
    file_exists = download_from_onedrive()

    # STOP if file not downloaded
    if not file_exists:
        raise Exception(
            "STOP: Excel file not downloaded. Preventing overwrite."
        )

    # Read existing Excel
    try:
        old_df = pd.read_excel(LOCAL_FILE)

        print(f"Existing rows     : {len(old_df)}")

        for col_check in [
            "Original Comment",
            "Complaint_Text",
            "text"
        ]:
            if col_check in old_df.columns:
                old_texts = set(
                    old_df[col_check]
                    .dropna()
                    .apply(clean_text)
                )
                break
        else:
            old_texts = set()

        print(
            f"Existing comments : "
            f"{len(old_texts)} fingerprints loaded"
        )

    except Exception as e:
        print(f"ERROR reading Excel: {e}")
        raise Exception(
            "STOP: Failed to read existing Excel file"
        )

    next_sno = get_next_sno(old_df)

    # 2 — Scrape
    print(f"\n[2/7] Scraping YouTube "
          f"({len(SEARCH_QUERIES)} targeted queries)...")

    all_comments = []
    seen_videos = set()
    seen_clean = set()

    for i, query in enumerate(SEARCH_QUERIES):

        print(f"  [{i+1}/{len(SEARCH_QUERIES)}] {query}")

        videos = search_videos(
            query,
            max_results=5
        )

        new_vids = [
            v for v in videos
            if v["video_id"] not in seen_videos
        ]

        for video in new_vids:

            seen_videos.add(video["video_id"])

            comments = get_comments(video)

            filtered_comments = []

            for c in comments:

                clean = clean_text(
                    c.get("text", "")
                )

                if clean in old_texts:
                    continue

                if clean in seen_clean:
                    continue

                seen_clean.add(clean)

                filtered_comments.append(c)

            comments = filtered_comments

            if comments:

                all_comments.extend(comments)

                print(
                    f"    → {video['title'][:42]} | "
                    f"{len(comments)} matched"
                )

            time.sleep(
                random.uniform(0.5, 1.0)
            )
    print(
        f"\nKeyword-matched: {len(all_comments)} "
        f"from {len(seen_videos)} videos"
    )

    if not all_comments:
        print("No comments found.")
        save_last_run()
        return

    # 3 — Filter
    print("\n[3/7] Filtering for quality complaints...")

    filtered = []
    rejected = {}

    seen_clean_filter = set()

    for c in all_comments:

        passed, reason = passes_filter(
            c,
            seen_clean_filter,
            old_texts
        )

        if passed:

            seen_clean_filter.add(
                clean_text(c.get("text", ""))
            )

            filtered.append(c)

        else:

            rejected[reason] = (
                rejected.get(reason, 0) + 1
            )

    print(f"Before : {len(all_comments)}")
    print(f"After  : {len(filtered)} quality complaints")

    for reason, count in sorted(
        rejected.items(),
        key=lambda x: -x[1]
    ):
        print(
            f"  {reason:<25}: {count} rejected"
        )

    if not filtered:
        print("Nothing passed filter. Exiting.")
        save_last_run()
        return

    # 4 — Select top complaints
    print(
        f"\n[4/7] Selecting top {MAX_TOTAL} "
        f"most relevant..."
    )

    top_comments = select_top_10(filtered)

    if not top_comments:
        print("No comments selected.")
        save_last_run()
        return

    # 5 — Gemini + VADER
    total_gemini_time = len(top_comments) * 6

    print(
        f"\n[5/7] Gemini analysis "
        f"({len(top_comments)} comments)..."
    )

    print(
        f"  Estimated time: "
        f"~{total_gemini_time}s "
        f"(5s gap per call to avoid rate limit)"
    )

    new_rows = []

    for i, c in enumerate(top_comments):

        print(
            f"\n  [{i+1}/{len(top_comments)}] "
            f"score={c.get('_score', 0)} | "
            f"{c.get('keyword_group', '')} | "
            f"{c.get('text', '')[:60]}..."
        )

        analysis = call_gemini_safe(c)

        if analysis is None:

            print(
                "  → Gemini failed 3 times, "
                "using raw comment as fallback"
            )

            analysis = fallback(c)

        elif not analysis.get("is_useful", True):

            print(
                "  → Gemini: not a quality "
                "complaint → skipped"
            )

            continue

        sentiment_label, vader_score = get_sentiment(
            c["text"]
        )

        group = c.get(
            "keyword_group",
            "Other"
        )

        owner, email = OWNERS.get(
            group,
            OWNERS["Other"]
        )

        new_rows.append({

            "S.No": next_sno,

            "System / Technology":
                analysis.get(
                    "system_technology",
                    group
                ),

            "FN1 / FN2":
                analysis.get(
                    "fn_type",
                    "Not specified"
                ),

            "Month":
                analysis.get(
                    "month",
                    "Not specified"
                ),

            "Year":
                analysis.get(
                    "year",
                    "Not specified"
                ),

            "Model":
                analysis.get(
                    "model",
                    "Not specified"
                ),

            "Defect / Feedback Summary":
                analysis.get(
                    "defect_summary",
                    c["text"][:150]
                ),

            "Cause":
                analysis.get(
                    "cause",
                    "Not specified"
                ),

            "Action":
                analysis.get(
                    "action",
                    "Not specified"
                ),

            "Sentiment":
                sentiment_label,

            "VADER Score":
                vader_score,

            "Owner Name":
                owner,

            "Owner Email":
                email,

            "Status":
                "Open",

            "Source":
                "YouTube",

            "Video URL":
                c.get("video_url", ""),

            "Date":
                c.get("date", ""),

            "Original Comment":
                c.get("text", ""),
        })

        next_sno += 1

        print(
            f"  → {sentiment_label} | "
            f"{analysis.get('fn_type', '?')} | "
            f"{analysis.get('model', '?')[:20]} | "
            f"{analysis.get('defect_summary', '?')[:45]}"
        )

    print(
        f"\nNew rows created: "
        f"{len(new_rows)}/{MAX_TOTAL}"
    )

    if not new_rows:
        print("No new rows.")
        save_last_run()
        return

    # 6 — Merge
    print("\n[6/7] Merging with existing data...")

    df_new = pd.DataFrame(new_rows)

    new_start = len(old_df)

    if old_df is not None and not old_df.empty:

        for col in EXCEL_COLUMNS:

            if col not in old_df.columns:
                old_df[col] = ""

        final_df = pd.concat(
            [
                old_df[EXCEL_COLUMNS],
                df_new[EXCEL_COLUMNS]
            ],
            ignore_index=True
        )

    else:

        final_df = df_new[
            EXCEL_COLUMNS
        ].copy()

    # Safety dedup
    final_df["_d"] = final_df[
        "Original Comment"
    ].apply(clean_text)

    before = len(final_df)

    final_df.drop_duplicates(
        subset=["_d"],
        inplace=True
    )

    final_df.drop(
        columns=["_d"],
        inplace=True
    )

    final_df.reset_index(
        drop=True,
        inplace=True
    )

    if before > len(final_df):

        print(
            f"Safety dedup removed "
            f"{before - len(final_df)} rows"
        )

    final_df["S.No"] = range(
        1,
        len(final_df) + 1
    )

    print(
        f"Total: {len(final_df)} rows "
        f"({new_start} existing + "
        f"{len(new_rows)} new)"
    )

    # 7 — Save + Upload
    print(
        "\n[7/7] Saving and uploading "
        "to Google Drive..."
    )

    save_excel(final_df, new_start)

    upload_to_onedrive()

    save_last_run()

    print("\n" + "=" * 60)
    print("PIPELINE COMPLETE")
    print(
        f"Total rows in Excel : "
        f"{len(final_df)}"
    )
    print(
        f"New rows this run   : "
        f"{len(new_rows)}"
    )
    print("=" * 60)


if __name__ == "__main__":
   
    main()
